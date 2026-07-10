"""
research_engine/exporters/excel_exporter.py
Stage 10 — Export Engine

Writes a Dataset and its analysis results to a formatted multi-sheet
Excel workbook (.xlsx).

Sheets produced
---------------
1. Raw Dataset         — all 58+ variables, one row per respondent
2. Demographics        — demographic variables only
3. Questionnaire Data  — Likert items + section means + satisfaction category
4. Observation Data    — facility checklist items
5. Descriptive Stats   — mean, SD, min, max per Likert item (Chapter Four table)
6. Frequency Tables    — one block per categorical variable
7. Crosstabulations    — one block per crosstab (with chi-square stats)
8. Codebook            — variable dictionary
9. Validation Report   — check results from the validator

Public API
----------
    export(dataset, questionnaire, variable_dictionary, validation_report,
           output_dir, study_title, seed, spss_maps)
    → Path  (path to the written .xlsx file)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from research_engine.models import Dataset, Questionnaire, VariableDictionary
from research_engine.validators.dataset_validator import ValidationReport
from research_engine.analysis import (
    frequency_table, all_categorical,
    describe_many, likert_summary,
    crosstab,
)

# ── Colour palette ─────────────────────────────────────────────
_NAVY   = "1F3864"
_TEAL   = "17375E"
_PURPLE = "4B0082"
_GOLD   = "FFC000"
_LTBLUE = "DCE6F1"
_LTGREY = "F2F2F2"
_WHITE  = "FFFFFF"
_GREEN  = "375623"
_ORANGE = "974706"
_RED    = "C00000"
_LGOLD  = "FFF2CC"


def _hdr(ws, row_num: int, bg: str = _NAVY, fg: str = _WHITE) -> None:
    fill   = PatternFill("solid", fgColor=bg)
    font   = Font(bold=True, color=fg, size=10)
    border = Border(
        bottom=Side(style="medium", color=_GOLD),
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )
    for cell in ws[row_num]:
        if cell.value is not None or True:
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border


def _alt(ws, start: int, end: int, n_cols: int) -> None:
    fe = PatternFill("solid", fgColor=_LTBLUE)
    fo = PatternFill("solid", fgColor=_LTGREY)
    for r in range(start, end + 1):
        fill = fe if r % 2 == 0 else fo
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = fill


def _autofit(ws, mn: int = 8, mx: int = 45) -> None:
    for col in ws.columns:
        w = max((len(str(cell.value or "")) for cell in col), default=mn)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)


def _section_hdr(ws, title: str, n_cols: int = 6, bg: str = _TEAL) -> None:
    ws.append([title] + [""] * (n_cols - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=_WHITE, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ── Public API ─────────────────────────────────────────────────

def export(
    dataset:              Dataset,
    questionnaire:        Questionnaire,
    variable_dictionary:  VariableDictionary,
    validation_report:    ValidationReport,
    output_dir:           str | Path,
    study_title:          str = "",
    seed:                 int | None = None,
    spss_maps:            dict | None = None,
    categorical_vars:     list[str] | None = None,
    crosstab_pairs:       list[tuple[str, str]] | None = None,
) -> Path:
    """
    Write the complete Excel workbook and return the file path.

    Parameters
    ----------
    dataset              : the generated Dataset
    questionnaire        : for Likert summary and structure
    variable_dictionary  : for codebook and labels
    validation_report    : from the validator
    output_dir           : directory to write the file
    study_title          : used in the filename
    seed                 : recorded in the workbook
    spss_maps            : {field: {label: code}} for SPSS sheet
    categorical_vars     : variable names to tabulate (defaults to auto-detect)
    crosstab_pairs       : [(row_var, col_var), ...] to include as crosstabs
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)

    vd      = variable_dictionary
    labels  = {v.name: v.label for v in vd}

    _sheet_raw(wb, dataset)
    _sheet_demographics(wb, dataset, vd)
    _sheet_questionnaire(wb, dataset, questionnaire)
    _sheet_observations(wb, dataset, vd)
    _sheet_descriptives(wb, dataset, questionnaire)
    _sheet_frequencies(wb, dataset, vd, categorical_vars)
    _sheet_crosstabs(wb, dataset, crosstab_pairs or _default_crosstab_pairs(vd))
    _sheet_codebook(wb, vd)
    _sheet_validation(wb, validation_report, study_title, seed)

    slug = study_title[:30].replace(" ", "_").replace("/", "-") if study_title else "study"
    xl_path = out / f"{slug}_{ts}.xlsx"
    wb.save(xl_path)
    return xl_path


# ── Sheet builders ─────────────────────────────────────────────

def _sheet_raw(wb, dataset: Dataset) -> None:
    ws = wb.create_sheet("Raw Dataset")
    ws.freeze_panes = "B2"
    records = dataset.to_records()
    if not records:
        return
    headers = list(records[0].keys())
    ws.append(headers)
    _hdr(ws, 1)
    for rec in records:
        ws.append([rec.get(h) for h in headers])
    _alt(ws, 2, len(records) + 1, len(headers))
    _autofit(ws)


def _sheet_demographics(wb, dataset: Dataset, vd: VariableDictionary) -> None:
    ws = wb.create_sheet("Demographics")
    ws.freeze_panes = "B2"
    demo_vars = [v.name for v in vd.by_section("demographics")]
    if not demo_vars:
        demo_vars = [
            k for k in dataset.variable_names
            if k not in ("respondent_id", "facility_id")
            and not k.startswith("s") and not k.startswith("mean_")
            and k not in ("overall_mean", "satisfaction_category", "obs_yes_count")
            and not k.startswith("ward")
        ]
    headers = ["respondent_id", "facility_id"] + demo_vars
    ws.append(headers)
    _hdr(ws, 1)
    for resp in dataset:
        row = [resp.respondent_id, resp.facility_id]
        row += [resp.demographics.get(v) for v in demo_vars]
        ws.append(row)
    _alt(ws, 2, len(dataset) + 1, len(headers))
    _autofit(ws)


def _sheet_questionnaire(wb, dataset: Dataset, questionnaire: Questionnaire) -> None:
    ws = wb.create_sheet("Questionnaire Data")
    ws.freeze_panes = "B2"
    q_vars   = [q.variable_name for q in questionnaire.all_questions]
    mean_vars = [f"mean_{s.key}" for s in questionnaire.sections]
    headers  = ["respondent_id"] + q_vars + mean_vars + ["overall_mean", "satisfaction_category"]
    ws.append(headers)
    _hdr(ws, 1)
    for resp in dataset:
        row = [resp.respondent_id]
        row += [resp.get_value(v) for v in q_vars + mean_vars + ["overall_mean", "satisfaction_category"]]
        ws.append(row)
    _alt(ws, 2, len(dataset) + 1, len(headers))
    _autofit(ws)


def _sheet_observations(wb, dataset: Dataset, vd: VariableDictionary) -> None:
    ws   = wb.create_sheet("Observation Data")
    obs_vars = [
        v.name for v in vd
        if v.section and v.section.startswith("observation")
        and v.name != "obs_yes_count"
    ]
    if not obs_vars:
        obs_vars = [
            k for k in dataset.variable_names
            if k not in ("respondent_id", "facility_id", "obs_yes_count")
            and not k.startswith("s") and not k.startswith("mean_")
            and k not in ("overall_mean", "satisfaction_category")
            and all(v.get_value(k) in ("Yes", "No", None)
                    for v in list(dataset)[:5])
        ]
    headers = ["respondent_id", "facility_id"] + obs_vars + ["obs_yes_count"]
    ws.append(headers)
    _hdr(ws, 1)
    for resp in dataset:
        row = [resp.respondent_id, resp.facility_id]
        row += [resp.observation_dict.get(v, resp.get_value(v)) for v in obs_vars]
        row += [resp.get_value("obs_yes_count")]
        ws.append(row)
    _alt(ws, 2, len(dataset) + 1, len(headers))
    _autofit(ws)


def _sheet_descriptives(wb, dataset: Dataset, questionnaire: Questionnaire) -> None:
    ws = wb.create_sheet("Descriptive Statistics")
    ws.column_dimensions["A"].width = 55
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 12

    summary = likert_summary(dataset, questionnaire)

    for section in questionnaire.sections:
        _section_hdr(ws, f"Section {section.key}: {section.title}", n_cols=8)
        ws.append(["Question", "Item Text", "N", "Mean", "SD", "Min", "Max", "Interpretation"])
        _hdr(ws, ws.max_row, _TEAL, _WHITE)

        items = summary.items_for_section(section.key)
        for item in items:
            ws.append([
                item.question_number, item.label, item.n,
                round(item.mean, 3), round(item.std, 3),
                1, 5, item.interpretation,
            ])

        sec_mean = summary.section_means.get(section.key, 0)
        ws.append([
            "", f"Section {section.key} Mean", "",
            round(sec_mean, 3), "", "", "", _likert_interp(sec_mean),
        ])
        ws.cell(ws.max_row, 2).font = Font(bold=True)
        ws.append([])

    ws.append(["", "OVERALL SATISFACTION MEAN", "",
               round(summary.overall_mean, 3), "", "", "",
               _likert_interp(summary.overall_mean)])
    ws.cell(ws.max_row, 2).font = Font(bold=True, color=_NAVY)


def _sheet_frequencies(wb, dataset: Dataset, vd: VariableDictionary,
                       categorical_vars: list[str] | None) -> None:
    ws = wb.create_sheet("Frequency Tables")
    ws.column_dimensions["A"].width = 38

    if categorical_vars is None:
        categorical_vars = _default_categorical_vars(vd, dataset)

    labels = {v.name: v.label for v in vd}
    tables = all_categorical(dataset, categorical_vars, labels=labels, sort_by="value")

    for table in tables:
        _section_hdr(ws, table.label, n_cols=4)
        ws.append(["Value", "Frequency", "Percent (%)", "Cumulative (%)"])
        _hdr(ws, ws.max_row, _TEAL, _WHITE)
        for row in table.to_rows():
            ws.append(list(row))
        ws.append([])

    _autofit(ws)


def _sheet_crosstabs(wb, dataset: Dataset, pairs: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet("Crosstabulations")
    ws.column_dimensions["A"].width = 32

    for row_var, col_var in pairs:
        try:
            ct = crosstab(dataset, row_var, col_var)
        except Exception as e:
            ws.append([f"Could not compute {row_var} × {col_var}: {e}"])
            ws.append([])
            continue

        _section_hdr(ws, f"{ct.row_label}  ×  {ct.col_label}", n_cols=len(ct.col_categories) + 2)
        for row in ct.to_rows():
            ws.append(list(row))
        ws.append([])

        # Stats block
        ws.append(["Chi-square (χ²)", "df", "p-value", "Cramer's V", "Significant?"])
        _hdr(ws, ws.max_row, _TEAL, _WHITE)
        ws.append(list(ct.stats_row()))
        if ct.note:
            ws.append([f"Note: {ct.note}"])
            ws.cell(ws.max_row, 1).font = Font(italic=True, color=_ORANGE)
        ws.append([])

    _autofit(ws)


def _sheet_codebook(wb, vd: VariableDictionary) -> None:
    ws = wb.create_sheet("Codebook")
    headers = ["Variable", "Label", "Type", "Scale", "Section",
               "Q Number", "Values / Range", "SPSS Codes", "Notes"]
    ws.append(headers)
    _hdr(ws, 1)
    for row in vd.to_codebook_rows():
        ws.append(list(row))
    _alt(ws, 2, len(list(vd)) + 1, len(headers))
    _autofit(ws)


def _sheet_validation(wb, report: ValidationReport,
                      study_title: str, seed: Any) -> None:
    ws = wb.create_sheet("Validation Report")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 75

    ws.append(["RESEARCH ANALYSIS TOOLKIT — VALIDATION REPORT"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(1, 1).font = Font(bold=True, size=13, color=_NAVY)

    ws.append([f"Study: {study_title}"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([f"Seed: {seed}"])
    ws.append([f"Result: {report.summary()}"])
    ws.append([])
    ws.append(["Status", "Check"])
    _hdr(ws, ws.max_row)

    colour = {"pass": _GREEN, "warn": _ORANGE, "error": _RED}
    label  = {"pass": "✓ PASS", "warn": "⚠ WARN", "error": "✗ ERR"}
    for check in report.checks:
        r = ws.max_row + 1
        ws.append([label[check.status], check.message])
        ws.cell(r, 1).font = Font(bold=True, color=colour[check.status])

    _autofit(ws)


# ── Helpers ────────────────────────────────────────────────────

def _default_categorical_vars(vd: VariableDictionary, dataset: Dataset) -> list[str]:
    from research_engine.models.variable import MeasurementScale
    cats = [
        v.name for v in vd
        if v.scale in (MeasurementScale.NOMINAL, MeasurementScale.ORDINAL)
        and not v.is_derived
        and v.name not in ("respondent_id",)
        and v.section == "demographics"
    ]
    cats.append("satisfaction_category")
    return cats


def _default_crosstab_pairs(vd: VariableDictionary) -> list[tuple[str, str]]:
    demo_cats = [
        v.name for v in vd
        if v.scale in ("Nominal", "Ordinal")
        and v.section == "demographics"
        and not v.is_derived
    ]
    from research_engine.models.variable import MeasurementScale
    demo_cats = [
        v.name for v in vd
        if v.scale in (MeasurementScale.NOMINAL, MeasurementScale.ORDINAL)
        and v.section == "demographics"
        and not v.is_derived
    ]
    return [(var, "satisfaction_category") for var in demo_cats[:4]]


def _likert_interp(mean: float) -> str:
    if mean >= 4.5: return "Very Satisfied"
    if mean >= 3.5: return "Satisfied"
    if mean >= 2.5: return "Neutral"
    if mean >= 1.5: return "Dissatisfied"
    return "Very Dissatisfied"
