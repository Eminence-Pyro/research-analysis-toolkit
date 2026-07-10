"""
research_engine/parsers/workbook_reader.py
Stage 2 — Readers (Input Layer)

Reads an Excel analysis workbook and extracts:
  - Variable names and labels from column headers
  - Section groupings from merged header rows or sheet names
  - Existing response data (if the workbook is pre-populated)

This allows a researcher to drop their existing analysis framework
Excel file into input/ and have the toolkit understand it automatically.

Public API
----------
    WorkbookReader(path)                          — initialise with file path
    WorkbookReader.read_variable_headers(sheet)   — extract variable names from a sheet
    WorkbookReader.read_framework_structure()      — detect section/variable layout
    WorkbookReader.sheet_names                     — list of sheet names

Note: Full implementation in Stage 3 (Variable Discovery Engine).
This file provides the structural foundation and header reader.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any


class WorkbookReader:
    """
    Reads an Excel workbook and returns domain-compatible structures.

    Uses openpyxl for reading. Does not depend on pandas — the toolkit
    controls how data is interpreted rather than delegating to DataFrame
    inference.

    Attributes
    ----------
    path : Path
        Path to the Excel file.
    """

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.path}")
        self._wb = None   # loaded lazily

    def _load(self):
        """Load the workbook (lazy — only when first accessed)."""
        if self._wb is None:
            try:
                import openpyxl
            except ImportError as exc:
                raise ImportError(
                    "openpyxl is required for WorkbookReader. "
                    "Install it with: pip install openpyxl"
                ) from exc
            self._wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        return self._wb

    @property
    def sheet_names(self) -> list[str]:
        """Names of all sheets in the workbook."""
        return self._load().sheetnames

    def read_variable_headers(self, sheet_name: str) -> list[str]:
        """
        Read the first non-empty row of a sheet as variable (column) headers.

        Returns a list of header strings with whitespace stripped and
        empty cells replaced with f"col_{n}".

        Parameters
        ----------
        sheet_name : str
            Name of the sheet to read.

        Returns
        -------
        list of str — one entry per column
        """
        wb = self._load()
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet {sheet_name!r} not found in workbook.")
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                return [
                    str(cell).strip() if cell is not None else f"col_{i}"
                    for i, cell in enumerate(row)
                ]
        return []

    def read_all_rows(
        self,
        sheet_name: str,
        skip_rows:  int = 1,
        max_rows:   int | None = None,
    ) -> list[dict[str, Any]]:
        """
        Read all data rows from a sheet as a list of dicts.

        The first non-empty row is treated as headers (auto-detected).
        Subsequent rows are zipped with those headers.

        Parameters
        ----------
        sheet_name : str
        skip_rows  : number of header rows to skip (default 1)
        max_rows   : stop after this many data rows (None = all)

        Returns
        -------
        list of dicts — one per data row
        """
        wb      = self._load()
        ws      = wb[sheet_name]
        rows    = list(ws.iter_rows(values_only=True))

        # Find first non-empty row as header
        header_idx = 0
        for i, row in enumerate(rows):
            if any(cell is not None for cell in row):
                header_idx = i
                break

        headers   = [
            str(c).strip() if c is not None else f"col_{j}"
            for j, c in enumerate(rows[header_idx])
        ]
        data_rows = rows[header_idx + skip_rows:]
        if max_rows:
            data_rows = data_rows[:max_rows]

        return [
            {headers[j]: cell for j, cell in enumerate(row)}
            for row in data_rows
            if any(cell is not None for cell in row)
        ]

    def read_framework_structure(self) -> dict[str, list[str]]:
        """
        Detect the variable structure of the workbook.

        Returns a dict mapping sheet_name → list of column headers
        for every sheet. Used by the Variable Discovery Engine (Stage 3)
        to infer sections and variables from an analysis framework.
        """
        return {name: self.read_variable_headers(name) for name in self.sheet_names}

    def close(self) -> None:
        """Release the file handle."""
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()

    def __repr__(self) -> str:
        return f"WorkbookReader(path={self.path.name!r}, sheets={len(self.sheet_names)})"
